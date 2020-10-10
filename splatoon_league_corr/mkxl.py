import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl.styles.fonts import Font
import sys
import pkgutil
import io
import pandas as pd
from . import team_data
import os

def outxlsx(df,save_dir,outname):
	df.to_csv(save_dir + '/' +outname+'.csv')
	data = pd.read_csv(save_dir + '/' +outname+'.csv')
	data.to_excel(save_dir + '/' +outname+'.xlsx', encoding='utf-8')
	os.remove(save_dir + '/' +outname+'.csv')

def outcorr(save_dir,gameslist,myteam):
	wb = openpyxl.Workbook()
	wslist = ['ガチエリア','ガチヤグラ','ガチホコ','ガチアサリ']
	for sheet_name in wslist:
		ws_copy = wb.create_sheet(title=sheet_name)
		ws = wb[sheet_name]
		ws['B2'].value = sheet_name + 'の相関係数表'
		ws['B2'].font = Font(size=24, italic=False)
		ws.cell(row=5, column=2,value='win').alignment = Alignment(horizontal='center',vertical = 'center',wrap_text=True)
		ws.cell(row=6, column=2,value='EnemyUdemae').alignment = Alignment(horizontal='center',vertical = 'center',wrap_text=True)
		ws.cell(row=3, column=4,value='win').alignment = Alignment(horizontal='center',vertical = 'center',wrap_text=True)
		ws.cell(row=3, column=5,value='EnemyUdemae').alignment = Alignment(horizontal='center',vertical = 'center',wrap_text=True)
		
		for i in range(4):
			ws['C'+str(7+4*i)].value = 'Kill/min'
			ws['C'+str(8+4*i)].value = 'Death/min'
			ws['C'+str(9+4*i)].value = 'PaintPoint/min'
			ws['C'+str(10+4*i)].value = 'Special/min'
		for i in range(4):
			ws.cell(row=4, column=6+4*i,value='Kill/min').alignment = Alignment(horizontal='center',vertical = 'center',wrap_text=True)
			ws.cell(row=4, column=7+4*i,value='Death/min').alignment = Alignment(horizontal='center',vertical = 'center',wrap_text=True)
			ws.cell(row=4, column=8+4*i,value='PaintPoint/min').alignment = Alignment(horizontal='center',vertical = 'center',wrap_text=True)
			ws.cell(row=4, column=9+4*i,value='Special/min').alignment = Alignment(horizontal='center',vertical = 'center',wrap_text=True)

		ws.cell(row=3, column=6,value=myteam.player.jpname).alignment = Alignment(horizontal='center',vertical = 'center',wrap_text=True)
		ws.cell(row=3, column=10,value=myteam.friend1.jpname).alignment = Alignment(horizontal='center',vertical = 'center',wrap_text=True)
		ws.cell(row=3, column=14,value=myteam.friend2.jpname).alignment = Alignment(horizontal='center',vertical = 'center',wrap_text=True)
		ws.cell(row=3, column=18,value=myteam.friend3.jpname).alignment = Alignment(horizontal='center',vertical = 'center',wrap_text=True)
		ws.cell(row=7, column=2,value=myteam.player.jpname).alignment = Alignment(horizontal='center',vertical = 'center',wrap_text=True)
		ws.cell(row=11, column=2,value=myteam.friend1.jpname).alignment = Alignment(horizontal='center',vertical = 'center',wrap_text=True)
		ws.cell(row=15, column=2,value=myteam.friend2.jpname).alignment = Alignment(horizontal='center',vertical = 'center',wrap_text=True)
		ws.cell(row=19, column=2,value=myteam.friend3.jpname).alignment = Alignment(horizontal='center',vertical = 'center',wrap_text=True)

		ws.merge_cells('B3:C4')
		
		ws.merge_cells('B5:C5')
		ws.merge_cells('B6:C6')
		ws.merge_cells('D3:D4')
		ws.merge_cells('E3:E4')
		
		ws.merge_cells('F3:I3')
		ws.merge_cells('J3:M3')
		ws.merge_cells('N3:Q3')
		ws.merge_cells('R3:U3')
		ws.merge_cells('B7:B10')
		ws.merge_cells('B11:B14')
		ws.merge_cells('B15:B18')
		ws.merge_cells('B19:B22')

		border = Border(top=Side(style='thin', color='000000'), 
						bottom=Side(style='thin', color='000000'), 
						left=Side(style='thin', color='000000'),
						right=Side(style='thin', color='000000')
		)
		for row_num in range(3,23):
			for col_num in range(2,22):
				ws.cell(row=row_num ,column=col_num).border = border
			
		for col_num in range(2,22):
			ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15

		for row_num in range(5,23):
			for col_num in range(4,22):
				if row_num==col_num+1:
					continue
				ws.cell(row=row_num ,column=col_num).number_format = '0.00'	

		ws.row_dimensions[2].height = 30

	xlsxlist = ['splat_zones_corr.xlsx','tower_control_corr.xlsx','rainmaker_corr.xlsx','clam_blitz_corr.xlsx']
	for k in range(4):
		corrwb = openpyxl.load_workbook(save_dir+'/'+xlsxlist[k])
		corrws = corrwb.active
		for i in range(1,19):
			for j in range(1,19):
				copy = corrws.cell(row = i+1, column =j+2).value
				wb[wslist[k]].cell(row = i+4, column = j+3 , value = copy)
		for i in range(1,19):
			wb[wslist[k]].cell(row = i+4, column = i+3 , value = "--").alignment = Alignment(horizontal='center',vertical = 'center',wrap_text=True)
		wb[wslist[k]]['K2'].value=gameslist[k]
		wb[wslist[k]]['L2'].value='試合'
		wb[wslist[k]]['K2'].font = Font(size=24, italic=False)
		wb[wslist[k]]['L2'].font = Font(size=24, italic=False)
		os.remove(save_dir+'/'+xlsxlist[k])
	
	wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
	wb.save(save_dir+'/result.xlsx')
	print("output : "+save_dir+'/league_all.xlsx')
	print("output : "+save_dir+'/result.xlsx')
